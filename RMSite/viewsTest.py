from django.shortcuts import render
from django.http import HttpResponse
from HR.models import Person

def testImport(request):
    return HttpResponse("该功能关闭")
    # import_Excel()
    # return HttpResponse("import data finlished.")

from openpyxl import load_workbook
def import_Excel():
    wb = load_workbook(r"E:\OneDrive\Work\billjc\区域研发中心\人力资源\在职&离职名单\南京在职&amp;离职名单--2018.5.11.xlsx",data_only=True)
    sht = wb["DU1-DU5"]
    # colIDX_SRC=["A","B"]
    # colIDX_DES=["A","B"]
    # 花名册Excel表格中，列名称对应的数据库字段属性，key是列名，value是字段属性
    colIDX_Fields_Mapping = {
        "B":"workNo",
        "C":"userName",
        "E":"IDNo",
        "F":"sex", #性别字段
        "G":"entryDate", # 到岗日期/入职日期
        "M":"depart",
        "P":"productUnit",
        "Q":"projectName",
        "V":"mobilePhone",
        "X":"email",
        "Z":"workNoExt",
        "AA":"emailExt",
        "AC":"provinceBirth",
        "AD":"cityBirth",
        "AG":"birthDay",
        "AH":"maritalStatus",
        "AJ":"graduatedSchool", #毕业院校
        "AM":"education", #学历
        "AO":"profession",  # 大学专业名称
        "AR":"graduatedDay", # 毕业时间
        "AW":"address",
        }
    usedRowCount = sht.max_row
    # usedRowCount = 10
    #数据从第二行开始读取，第一行是标题
    for row in range(2,usedRowCount):
        p = Person()
        for colName,fieldName  in colIDX_Fields_Mapping.items():
            print(colName,fieldName)
            #动态设置对象的属性值
            p.__setattr__ (fieldName, sht["%s%i" % (colName,row)].value)
            # pars[fieldName]=sht["%s%i" % (colName,row)].value
        p.save()
